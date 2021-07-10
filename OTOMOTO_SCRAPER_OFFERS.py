from tkinter import *
from tkinter import messagebox
from csv import reader
from csv import writer
import csv
import pandas as pd
import os 
import sys
from bs4 import BeautifulSoup
import requests
import openpyxl as exl
from openpyxl import styles
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, Fill

filename = 'LIST.csv'

# READ CSV AND INPUT TO LISTBOX
with open(filename, 'r') as csv_file:
    csv_reader = reader(csv_file)
    task_list = list(csv_reader)    

def zmien_format(cena):
            return (cena.replace(' ','').replace(',','').replace('PLN',' PLN'))

# ADD TASKS
def newTask():
    task = my_entry.get()
    if task != "":
        lb.insert(END, task)
        with open(filename,'a+', newline='') as fd:
            fd.write(task + '\n')
        my_entry.delete(0, "end")
    else:
        messagebox.showwarning("warning", "Please enter some task.")

# DELETE TASKS
def deleteTask():
    
    lb.delete(ANCHOR)
    
    with open(filename,"rb") as source:
        rdr= csv.reader(source)
        with open(filename,"wb") as result:
            wtr= csv.writer(result)
            for r in rdr:
                wtr.writerow(r[0])
           
# DOWNLOAD PICTURE AND DATA FROM OTOMOTO.PL
def download():

    for Linki in list(task_list):
        odpowiedz = requests.get(', '.join(Linki))
        soup = BeautifulSoup(odpowiedz.text, 'html.parser')

        tytul = soup.find('span', class_='offer-title big-text fake-title').text.strip()
        cena = soup.find('span', class_='offer-price__number').text
        
        count=len(soup.find_all('li', class_='offer-params__item'))+1

        foto = soup.find('img', class_='bigImage')        
        response = requests.get(foto['data-lazy'])
        file = open(str(tytul) + ".png", "wb")
        file.write(response.content)
        file.close()
        
        # GENERATE OFFERS IN EXCEL
        wb = exl.Workbook()
        ws = wb.worksheets[0]
        img = exl.drawing.image.Image(str(tytul) + ".png")
        img.anchor = 'A'+ str(count)
        ws.add_image(img)
        
        # INPUT DATA TO CELLS
        ws['A1'].value = tytul + '  |  cena: ' + zmien_format(cena)
        ws.merge_cells('A1:Q1')
        ws['A2'].value = "SPECYFIKACJA:"
        ws['G2'].value = "WYPOSAŻENIE:"

        i = 0
        
        for item in soup.find_all('li', class_='offer-params__item'):
            kol = item.find('span', class_='offer-params__label').text.strip()
            value = item.find('div', class_='offer-params__value').text.strip()        
            info = kol + ': ' + value
            i += 1
            ws['A'+ str(2+i)].value = info
            ws['A'+ str(2+i)].font= Font(color="000000", size=14, bold=True)
        
        o = 0

        for item_wyp in soup.find_all('li', class_='offer-features__item'):                   
            wypo = item_wyp.find('span').text
            o += 1
            ws['G'+ str(2+o)].value = item_wyp.text.strip()
            ws['G'+ str(2+o)].font= Font(color="000000", size=14, bold=True)
        
        # FORMATTING
        ws['A1'].font= Font(color="FF0000", size=28, bold=True)
        ws['A2'].font= Font(color="FF0000", size=20, bold=True)
        ws['G2'].font= Font(color="FF0000", size=20, bold=True)
        
        wb.save(str(tytul) + '.xlsx')
    
# CREATE GUI INTERFACE FOR INPUT VALUES 
ws = Tk()
ws.geometry('900x600+400+200')
ws.title('LISTA')
ws.config(bg='#223441')
ws.resizable(width=False, height=False)

frame = Frame(ws)
frame.pack(pady=10)

lb = Listbox(
    frame,
    width=70,
    height=8,
    font=('Times', 18),
    bd=0,
    fg='#464646',
    highlightthickness=0,
    selectbackground='#a6a6a6',
    activestyle="none",
    
)
lb.pack(side=LEFT, fill=BOTH)


for item in task_list:
    lb.insert(END, item)

sb = Scrollbar(frame)
sb.pack(side=RIGHT, fill=BOTH)

lb.config(yscrollcommand=sb.set)
sb.config(command=lb.yview)

my_entry = Entry(
    ws,
    font=('times', 24)
    )

my_entry.pack(pady=20)

button_frame = Frame(ws)
button_frame.pack(pady=20)

addTask_btn = Button(
    button_frame,
    text='DODAJ',
    font=('times 14'),
    bg='#c5f776',
    padx=20,
    pady=10,
    command=newTask
)
addTask_btn.pack(fill=BOTH, expand=True, side=LEFT)

delTask_btn = Button(
    button_frame,
    text='USUŃ',
    font=('times 14'),
    bg='#ff8b61',
    padx=20,
    pady=10,
    command=deleteTask
)
delTask_btn.pack(fill=BOTH, expand=True, side=LEFT)

goTask_btn = Button(
    button_frame,
    text='START',
    font=('times 14'),
    bg='green',
    padx=20,
    pady=10,
    command=download
)
goTask_btn.pack(fill=BOTH, expand=True, side=BOTTOM)

ws.mainloop()
